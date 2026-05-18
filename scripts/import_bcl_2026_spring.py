"""Import the inaugural Belgian Carcassonne League season (2026 Spring).

Reads `data/raw/bcl/BELGIUM.xlsx` (four sheets: ML/GL/SL/BL) and inserts:
  - 4 tournaments (one per league tier) of type 'BCL'
  - tournament_matches for every played duel (incl. Cross Final in ML)
  - tournament_participants with final_rank derived from Stage 1 standings,
    using the Cross Final to decide ranks 1-2 in ML

Best-of-3 format (Bo3): score_1/score_2 are game-win counts (0/1/2).
Per-game point scores are not imported here — BGA table URLs are stored
in tournament_matches.notes for a future hydration script.
"""
from __future__ import annotations

from pathlib import Path

import duckdb
from openpyxl import load_workbook

ROOT = Path(__file__).parents[1]
DB_PATH = ROOT / "data" / "carcassonne.duckdb"
EXCEL = ROOT / "data" / "raw" / "bcl" / "BELGIUM.xlsx"

SEASON_NAME = "BCL 2026 Spring"
SEASON_YEAR = 2026
SEASON_EDITION = "2026 Spring"
SEASON_DATE_START = "2026-03-02"
SEASON_DATE_END = "2026-05-17"

LEAGUES = [
    # (tier, sheet name, full label)
    ("ML", "BCL-2026S-ML", "Master League"),
    ("GL", "BCL-2026S-GL", "Gold League"),
    ("SL", "BCL-2026S-SL", "Silver League"),
    ("BL", "BCL-2026S-BL", "Bronze League"),
]

# Column indices (1-based, matching openpyxl ws.cell()):
COL_ID = 2
COL_STAGE = 5
COL_GROUP = 6
COL_ROUND = 8
COL_P1_BGA = 16
COL_P1_NAME = 17
COL_GW1 = 18
COL_GW2 = 19
COL_P2_NAME = 20
COL_P2_BGA = 21
COL_STATUS = 23
COL_LINKS = 27

# Alias map: BGA handle (excel) -> alternate names that may exist in DB.
NAME_ALIASES = {
    "MarathonMeeple": "jorenderidder",
    "jeanleon": "zusjezus",
    "zusjezus": "jeanleon",
}


def find_or_create_player(con, name: str, bga_id: str | None, cache: dict) -> int:
    key = (name.lower(), bga_id or "")
    if key in cache:
        return cache[key]

    if bga_id:
        row = con.execute(
            "SELECT id FROM players WHERE bga_player_id = ?", [bga_id]
        ).fetchone()
        if row:
            cache[key] = row[0]
            return row[0]

    candidates = {name}
    if name in NAME_ALIASES:
        candidates.add(NAME_ALIASES[name])
    for cand in candidates:
        row = con.execute(
            "SELECT id FROM players "
            "WHERE LOWER(name) = LOWER(?) OR LOWER(name_nl) = LOWER(?)",
            [cand, cand],
        ).fetchone()
        if row:
            if bga_id:
                con.execute(
                    "UPDATE players SET bga_player_id = ? "
                    "WHERE id = ? AND (bga_player_id IS NULL OR bga_player_id = '')",
                    [bga_id, row[0]],
                )
            cache[key] = row[0]
            return row[0]

    new_id = con.execute(
        "INSERT INTO players (name, bga_player_id) VALUES (?, ?) RETURNING id",
        [name, bga_id],
    ).fetchone()[0]
    print(f"  created player {new_id}: {name} (bga={bga_id})")
    cache[key] = new_id
    return new_id


def parse_int(value) -> int | None:
    if value is None or value == "" or value == "-":
        return None
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return None


def parse_sheet(ws):
    """Yield dicts of valid match rows; skip ERROR/empty/non-numeric score rows."""
    for r in range(2, ws.max_row + 1):
        rid = ws.cell(r, COL_ID).value
        if not rid:
            continue
        status = ws.cell(r, COL_STATUS).value
        gw1 = parse_int(ws.cell(r, COL_GW1).value)
        gw2 = parse_int(ws.cell(r, COL_GW2).value)
        if gw1 is None or gw2 is None:
            print(f"    skip row {r}: non-numeric GW (status={status})")
            continue
        yield {
            "row": r,
            "stage": ws.cell(r, COL_STAGE).value,
            "group": ws.cell(r, COL_GROUP).value,
            "round": ws.cell(r, COL_ROUND).value,
            "p1_bga": str(ws.cell(r, COL_P1_BGA).value or "") or None,
            "p1_name": ws.cell(r, COL_P1_NAME).value,
            "gw1": gw1,
            "gw2": gw2,
            "p2_name": ws.cell(r, COL_P2_NAME).value,
            "p2_bga": str(ws.cell(r, COL_P2_BGA).value or "") or None,
            "links": ws.cell(r, COL_LINKS).value,
        }


def stage_label(tier: str, match: dict) -> str:
    if match["stage"] == "Stage 2":
        return match["round"] or "Stage 2"
    if tier == "ML":
        grp = match["group"]
        return f"Group {grp}" if grp else "Round-robin"
    return "Round-robin"


def round_short(round_text: str | None) -> str | None:
    if not round_text:
        return None
    s = str(round_text)
    # "Round 1: 02/03 - 15/03" -> "Round 1"
    if ":" in s:
        return s.split(":", 1)[0].strip()
    return s.strip()


def compose_notes(round_text: str | None, links: str | None) -> str | None:
    parts = []
    rn = round_short(round_text)
    if rn:
        parts.append(rn)
    if links:
        parts.append(f"links={links}")
    return " | ".join(parts) if parts else None


def next_tournament_id(con) -> int:
    row = con.execute("SELECT COALESCE(MAX(id), 0) + 1 FROM tournaments").fetchone()
    return int(row[0])


def main() -> None:
    if not EXCEL.exists():
        raise SystemExit(f"Excel not found: {EXCEL}")
    con = duckdb.connect(str(DB_PATH))

    # Reserve 4 contiguous tournament ids.
    tid_base = next_tournament_id(con)
    tids = {tier: tid_base + i for i, (tier, _, _) in enumerate(LEAGUES)}

    # Bail if any of those ids exist (re-running safely).
    for tid in tids.values():
        existing = con.execute(
            "SELECT id, name FROM tournaments WHERE id = ?", [tid]
        ).fetchone()
        if existing:
            con.close()
            raise SystemExit(
                f"Tournament id={existing[0]} ({existing[1]!r}) already exists; aborting."
            )

    # Refuse to double-import the same season.
    dup = con.execute(
        "SELECT id, name FROM tournaments WHERE type='BCL' AND name LIKE ?",
        [f"{SEASON_NAME}%"],
    ).fetchone()
    if dup:
        con.close()
        raise SystemExit(
            f"A {SEASON_NAME} tournament already exists (id={dup[0]}, name={dup[1]!r}); aborting."
        )

    # Some sequences have drifted behind their table's max(id) due to past
    # manual inserts; burn them forward so new INSERTs don't collide.
    for table, seq in [
        ("tournament_matches", "tournament_matches_id_seq"),
        ("tournament_participants", "tournament_participants_id_seq"),
    ]:
        max_id = con.execute(
            f"SELECT COALESCE(MAX(id), 0) FROM {table}"
        ).fetchone()[0]
        while True:
            nv = con.execute(f"SELECT nextval('{seq}')").fetchone()[0]
            if nv > max_id:
                break

    wb = load_workbook(EXCEL, data_only=True)
    player_cache: dict = {}
    summary = []

    for tier, sheet_name, full_label in LEAGUES:
        tid = tids[tier]
        name = f"{SEASON_NAME} {tier}"
        edition = f"{SEASON_EDITION} {tier}"
        notes = f"{full_label} — Season 1 (Spring 2026)."

        con.execute(
            """
            INSERT INTO tournaments (id, name, type, year, edition,
                                     date_start, date_end, notes)
            VALUES (?, ?, 'BCL', ?, ?, ?, ?, ?)
            """,
            [tid, name, SEASON_YEAR, edition,
             SEASON_DATE_START, SEASON_DATE_END, notes],
        )
        print(f"Created tournament {tid}: {name}")

        ws = wb[sheet_name]
        matches_inserted = 0
        player_ids: set[int] = set()
        for i, m in enumerate(parse_sheet(ws), start=1):
            p1_id = find_or_create_player(con, m["p1_name"], m["p1_bga"], player_cache)
            p2_id = find_or_create_player(con, m["p2_name"], m["p2_bga"], player_cache)
            player_ids.update([p1_id, p2_id])

            stage = stage_label(tier, m)
            s1, s2 = m["gw1"], m["gw2"]
            if s1 > s2:
                result = "1"
            elif s2 > s1:
                result = "2"
            else:
                result = "D"

            con.execute(
                """
                INSERT INTO tournament_matches
                    (tournament_id, stage, match_number, player_1_id, player_2_id,
                     score_1, score_2, result, notes)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                [tid, stage, i, p1_id, p2_id, s1, s2, result,
                 compose_notes(m["round"], m["links"])],
            )
            matches_inserted += 1

        print(f"  inserted {matches_inserted} matches "
              f"({len(player_ids)} unique players)")
        summary.append((tier, tid, matches_inserted, player_ids))

    # ── Compute participants + final_rank per league ──────────────────────────

    for tier, tid, _, _ in summary:
        compute_participants(con, tid, tier)

    # Set participants_count on each tournament row.
    for _, tid, _, _ in summary:
        n = con.execute(
            "SELECT COUNT(*) FROM tournament_participants WHERE tournament_id = ?",
            [tid],
        ).fetchone()[0]
        con.execute(
            "UPDATE tournaments SET participants_count = ? WHERE id = ?", [n, tid]
        )

    con.close()
    print(f"\nDone. Tournament ids {tid_base}..{tid_base + len(LEAGUES) - 1}.")


def compute_participants(con, tid: int, tier: str) -> None:
    """Build tournament_participants for one league tournament.

    Stage 1 round-robin determines per-group standings (W ↓, GDiff ↓, GW ↓).
    For ML, the Cross Final outcome decides ranks 1 and 2 (group winners);
    other ML players inherit their within-group rank tier.
    For GL/SL/BL there's a single round-robin → straight ranking.
    """
    rows = con.execute(
        """
        SELECT player_1_id, player_2_id, score_1, score_2, result, stage
        FROM tournament_matches
        WHERE tournament_id = ?
        """,
        [tid],
    ).fetchall()

    # Aggregate per-player stats from group-stage matches only.
    group_of: dict[int, str] = {}
    stats: dict[int, dict] = {}
    cross_final: tuple[int, int, str] | None = None  # (p1, p2, result)

    for p1, p2, s1, s2, res, stage in rows:
        if stage == "Cross Final":
            cross_final = (p1, p2, res)
            continue
        for pid, gw, gl, won in ((p1, s1, s2, res == "1"),
                                  (p2, s2, s1, res == "2")):
            st = stats.setdefault(pid, {"dp": 0, "dw": 0, "dl": 0,
                                          "gw": 0, "gl": 0})
            st["dp"] += 1
            st["gw"] += gw
            st["gl"] += gl
            if won:
                st["dw"] += 1
            elif res != "D":
                st["dl"] += 1
            # Stage 1 group label assignment (ML has Group A/B; others "Round-robin").
            group_of.setdefault(pid, stage)

    # Per-group ranking by (W desc, GDiff desc, GW desc).
    by_group: dict[str, list[int]] = {}
    for pid, st in stats.items():
        by_group.setdefault(group_of[pid], []).append(pid)
    for grp, pids in by_group.items():
        pids.sort(key=lambda p: (
            -stats[p]["dw"],
            -(stats[p]["gw"] - stats[p]["gl"]),
            -stats[p]["gw"],
        ))

    # Assemble final_rank.
    final_rank: dict[int, int] = {}

    if tier == "ML":
        groups = sorted(by_group.keys())  # e.g. ['Group A', 'Group B']
        # Group winners → ranks 1 & 2 via Cross Final, else by group sort order.
        winners = [by_group[g][0] for g in groups]
        if cross_final and cross_final[2] in ("1", "2"):
            p1, p2, res = cross_final
            cf_winner = p1 if res == "1" else p2
            cf_loser = p2 if res == "1" else p1
            final_rank[cf_winner] = 1
            final_rank[cf_loser] = 2
        else:
            for idx, w in enumerate(winners, start=1):
                final_rank[w] = idx
        # Remaining within-group positions interleave: 2nd-in-A → 3, 2nd-in-B → 4, etc.
        max_depth = max(len(pids) for pids in by_group.values())
        rank_cursor = 3
        for depth in range(1, max_depth):
            for g in groups:
                pids = by_group[g]
                if depth < len(pids):
                    final_rank[pids[depth]] = rank_cursor
                    rank_cursor += 1
    else:
        # Single round-robin → direct ranking.
        only_group = next(iter(by_group))
        for idx, pid in enumerate(by_group[only_group], start=1):
            final_rank[pid] = idx

    # Persist participants.
    for pid, st in stats.items():
        dp, dw, dl, gw, gl = st["dp"], st["dw"], st["dl"], st["gw"], st["gl"]
        win_pct = (dw / dp) if dp else 0.0
        resistance = ((gw - gl) / dp) if dp else 0.0
        con.execute(
            """
            INSERT INTO tournament_participants
                (tournament_id, player_id, final_rank,
                 duels_played, duels_won, duels_lost,
                 games_won, games_lost, win_pct, resistance)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            [tid, pid, final_rank.get(pid),
             dp, dw, dl, gw, gl,
             round(win_pct, 6), round(resistance, 6)],
        )

    print(f"  tournament {tid} ({tier}): {len(stats)} participants "
          f"ranked (cross_final={'yes' if cross_final else 'no'})")


if __name__ == "__main__":
    main()
