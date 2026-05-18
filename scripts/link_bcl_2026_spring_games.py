"""Hydrate tournament_matches.game_id_* for BCL 2026 Spring matches.

Reads the original Excel to recover per-match Time Start (UTC) + BGA player ids,
then searches `games`/`game_players` for clusters of BGA games between the same
two players and assigns up to 5 game_id_* columns on the matching
`tournament_matches` row.

Heuristic: cluster BGA games where consecutive ended within 30 min; accept
clusters of size between n_expected and n_expected + 2 (player abandoned a
game then replayed). Disambiguate by proximity to `Time Start, UTC` from Excel.

Idempotent: skips matches that already have game_id_1 set.
"""
from __future__ import annotations

from datetime import timedelta
from pathlib import Path

import duckdb
from openpyxl import load_workbook

ROOT = Path(__file__).parents[1]
DB_PATH = ROOT / "data" / "carcassonne.duckdb"
EXCEL = ROOT / "data" / "raw" / "bcl" / "BELGIUM.xlsx"

SEASON_LO = "2026-02-15"
SEASON_HI = "2026-05-31"
MAX_GAP_MINUTES = 30
MAX_EXTRA_GAMES = 2

# Map sheet -> tournament_id (from the import; verified post-import).
SHEET_TO_TID = {
    "BCL-2026S-ML": 119,
    "BCL-2026S-GL": 120,
    "BCL-2026S-SL": 121,
    "BCL-2026S-BL": 122,
}

# Column indices (1-based), matching the importer.
COL_ID = 2
COL_STAGE = 5
COL_GROUP = 6
COL_TS_START = 15
COL_P1_BGA = 16
COL_P1_NAME = 17
COL_GW1 = 18
COL_GW2 = 19
COL_P2_NAME = 20
COL_P2_BGA = 21


def cluster_games(games, max_gap_minutes=MAX_GAP_MINUTES):
    """Group games into time clusters (consecutive within max_gap_minutes)."""
    clusters: list[list] = []
    current: list = []
    for g in games:
        if current and (g[1] - current[-1][1]).total_seconds() > max_gap_minutes * 60:
            clusters.append(current)
            current = []
        current.append(g)
    if current:
        clusters.append(current)
    return clusters


def stage_label_from_excel(stage_cell, group_cell, tier: str) -> str:
    if stage_cell == "Stage 2":
        return "Cross Final" if tier == "ML" else "Stage 2"
    if tier == "ML":
        return f"Group {group_cell}" if group_cell else "Round-robin"
    return "Round-robin"


def main() -> None:
    if not EXCEL.exists():
        raise SystemExit(f"Excel not found: {EXCEL}")
    con = duckdb.connect(str(DB_PATH))
    wb = load_workbook(EXCEL, data_only=True)

    total_updated = 0
    total_skipped_done = 0
    total_unresolved = 0

    for sheet, tid in SHEET_TO_TID.items():
        ws = wb[sheet]
        tier = sheet.split("-")[-1]   # 'ML' / 'GL' / 'SL' / 'BL'

        updated = 0
        skipped = 0
        unresolved = 0

        for r in range(2, ws.max_row + 1):
            rid = ws.cell(r, COL_ID).value
            if not rid:
                continue
            gw1 = ws.cell(r, COL_GW1).value
            gw2 = ws.cell(r, COL_GW2).value
            if not (isinstance(gw1, (int, float)) and isinstance(gw2, (int, float))):
                continue
            n_expected = int(gw1) + int(gw2)
            if n_expected == 0:
                continue

            p1_bga = str(ws.cell(r, COL_P1_BGA).value or "")
            p2_bga = str(ws.cell(r, COL_P2_BGA).value or "")
            if not p1_bga or not p2_bga:
                continue

            p1 = con.execute(
                "SELECT id FROM players WHERE bga_player_id = ?", [p1_bga]
            ).fetchone()
            p2 = con.execute(
                "SELECT id FROM players WHERE bga_player_id = ?", [p2_bga]
            ).fetchone()
            if not p1 or not p2:
                unresolved += 1
                continue
            p1_id, p2_id = p1[0], p2[0]

            stage_label = stage_label_from_excel(
                ws.cell(r, COL_STAGE).value, ws.cell(r, COL_GROUP).value, tier
            )

            # Locate the matching tournament_matches row.
            tm_row = con.execute(
                """
                SELECT id, score_1, score_2, game_id_1
                FROM tournament_matches
                WHERE tournament_id = ?
                  AND stage = ?
                  AND ((player_1_id = ? AND player_2_id = ?)
                       OR (player_1_id = ? AND player_2_id = ?))
                  AND score_1 = ? AND score_2 = ?
                LIMIT 1
                """,
                [tid, stage_label,
                 p1_id, p2_id, p2_id, p1_id,
                 int(gw1), int(gw2)],
            ).fetchone()
            if not tm_row:
                # Try with swapped scores (rare ordering edge case).
                tm_row = con.execute(
                    """
                    SELECT id, score_1, score_2, game_id_1
                    FROM tournament_matches
                    WHERE tournament_id = ?
                      AND stage = ?
                      AND ((player_1_id = ? AND player_2_id = ?)
                           OR (player_1_id = ? AND player_2_id = ?))
                    LIMIT 1
                    """,
                    [tid, stage_label, p1_id, p2_id, p2_id, p1_id],
                ).fetchone()
                if not tm_row:
                    print(f"  {sheet} row {r}: no tournament_matches row for "
                          f"{ws.cell(r, COL_P1_NAME).value} vs {ws.cell(r, COL_P2_NAME).value} "
                          f"(stage={stage_label})")
                    unresolved += 1
                    continue
            tm_id, _, _, existing_g1 = tm_row
            if existing_g1 is not None:
                skipped += 1
                continue

            ts_start = ws.cell(r, COL_TS_START).value

            games = con.execute(
                """
                SELECT g.id, COALESCE(g.ended_at, g.played_at) AS ts
                FROM games g
                JOIN game_players gp1 ON gp1.game_id = g.id AND gp1.player_id = ?
                JOIN game_players gp2 ON gp2.game_id = g.id AND gp2.player_id = ?
                WHERE COALESCE(g.ended_at, g.played_at) BETWEEN ? AND ?
                  AND (g.boardgame_id IS NULL OR g.boardgame_id = 1)
                ORDER BY ts
                """,
                [p1_id, p2_id, SEASON_LO, SEASON_HI],
            ).fetchall()

            clusters = cluster_games(games)
            valid = [
                cl for cl in clusters
                if n_expected <= len(cl) <= n_expected + MAX_EXTRA_GAMES
            ]

            if not valid:
                unresolved += 1
                continue

            chosen = None
            if len(valid) == 1:
                chosen = valid[0]
            elif ts_start:
                chosen = min(valid,
                             key=lambda cl: abs((cl[0][1] - ts_start).total_seconds()))
            else:
                # Multiple clusters, no timestamp: pick the smallest (likely the
                # actual match, not a replay/extra session).
                chosen = min(valid, key=len)

            # Truncate to n_expected games (drop extra abandoned/replay games
            # past the win count — the first n_expected in the cluster are the
            # actual match games in BGA chronological order).
            chosen = chosen[:n_expected]
            game_ids = [g[0] for g in chosen] + [None] * (5 - len(chosen))

            con.execute(
                """
                UPDATE tournament_matches
                SET game_id_1 = ?, game_id_2 = ?, game_id_3 = ?,
                    game_id_4 = ?, game_id_5 = ?
                WHERE id = ?
                """,
                [*game_ids, tm_id],
            )
            updated += 1

        total_updated += updated
        total_skipped_done += skipped
        total_unresolved += unresolved
        print(f"{sheet} (tid {tid}): linked {updated}, "
              f"skipped {skipped} (already linked), unresolved {unresolved}")

    con.close()
    print(f"\nDone. Total linked: {total_updated}, "
          f"skipped: {total_skipped_done}, unresolved: {total_unresolved}.")


if __name__ == "__main__":
    main()
